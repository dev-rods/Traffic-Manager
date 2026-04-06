"""Import patients from xlsx into Clínica Essência Estética."""
import openpyxl
import psycopg2
import os
import re
import hashlib
from dotenv import load_dotenv

load_dotenv()


def main():
    wb = openpyxl.load_workbook("C:/Users/andre/Downloads/Patient (2).xlsx")
    ws = wb.active

    clinic_name = "Clínica Essência Estética"
    base = "".join(e for e in clinic_name if e.isalnum()).lower()
    hash_suffix = hashlib.md5(clinic_name.encode()).hexdigest()[:6]
    clinic_id = f"{base}-{hash_suffix}"
    print(f"clinic_id: {clinic_id}")

    conn = psycopg2.connect(
        host=os.getenv("RDS_HOST"),
        port=os.getenv("RDS_PORT"),
        database=os.getenv("RDS_DATABASE"),
        user=os.getenv("RDS_USERNAME"),
        password=os.getenv("RDS_PASSWORD"),
    )
    cur = conn.cursor()

    inserted = 0
    skipped = 0

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        (active, age, birth, deleted, email, how, imp_id, ins_date,
         phone, name, notes, doc, other_phones, person, sex, typ, pid) = row

        if deleted or active != "X" or not phone:
            skipped += 1
            continue

        # Normalize phone: digits only, ensure 55 prefix
        phone_digits = re.sub(r"\D", "", str(phone))
        if len(phone_digits) <= 11:
            phone_digits = "55" + phone_digits

        gender = None
        if sex == "F":
            gender = "F"
        elif sex == "M":
            gender = "M"

        cur.execute(
            """
            INSERT INTO scheduler.patients (clinic_id, phone, name, gender)
            VALUES (%s, %s, %s, %s)
            ON CONFLICT (clinic_id, phone) DO UPDATE SET
                name=EXCLUDED.name, gender=EXCLUDED.gender, updated_at=NOW()
            """,
            (clinic_id, phone_digits, name, gender),
        )
        inserted += 1

    conn.commit()
    cur.close()
    conn.close()

    print(f"Inserted/Updated: {inserted}")
    print(f"Skipped: {skipped}")


if __name__ == "__main__":
    main()
