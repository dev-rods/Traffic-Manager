"""Import appointments from xlsx into Clínica Essência Estética."""
import openpyxl
import psycopg2
import os
import re
from datetime import datetime
from dotenv import load_dotenv

load_dotenv()

CLINIC_ID = "clinicaessenciaestetica-9668a4"
TARGET_DATES = {"2026-04-15", "2026-04-28", "2026-04-29"}


def normalize_phone(phone):
    if not phone:
        return None
    digits = re.sub(r"\D", "", str(phone))
    if len(digits) <= 11:
        digits = "55" + digits
    return digits


def main():
    wb = openpyxl.load_workbook("C:/Users/andre/Downloads/Appointment.xlsx")
    ws = wb.active

    conn = psycopg2.connect(
        host=os.getenv("RDS_HOST"),
        port=os.getenv("RDS_PORT"),
        database=os.getenv("RDS_DATABASE"),
        user=os.getenv("RDS_USERNAME"),
        password=os.getenv("RDS_PASSWORD"),
    )
    cur = conn.cursor()

    # Get service_id for Depilação a Laser
    cur.execute(
        "SELECT id FROM scheduler.services WHERE clinic_id = %s AND name LIKE '%%Depila%%' LIMIT 1",
        (CLINIC_ID,),
    )
    service_id = cur.fetchone()[0]
    print(f"Service ID: {service_id}")

    inserted = 0
    skipped = 0
    errors = []

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        (cat, create, deleted, dentId, dentName, ins, phone,
         notes, patId, patName, procs, status, dateStr, fromT, apptId, toT) = row

        if deleted:
            skipped += 1
            continue

        if not dateStr:
            skipped += 1
            continue

        # Parse date
        dt = datetime.fromisoformat(dateStr.replace("Z", "+00:00"))
        date_key = dt.strftime("%Y-%m-%d")

        if date_key not in TARGET_DATES:
            skipped += 1
            continue

        # Find patient by phone
        phone_norm = normalize_phone(phone)
        patient_id = None
        if phone_norm:
            cur.execute(
                "SELECT id FROM scheduler.patients WHERE clinic_id = %s AND phone = %s",
                (CLINIC_ID, phone_norm),
            )
            pat_row = cur.fetchone()
            if pat_row:
                patient_id = pat_row[0]

        # Calculate duration
        from_parts = fromT.split(":")
        to_parts = toT.split(":")
        from_min = int(from_parts[0]) * 60 + int(from_parts[1])
        to_min = int(to_parts[0]) * 60 + int(to_parts[1])
        duration = to_min - from_min

        try:
            cur.execute(
                """
                INSERT INTO scheduler.appointments
                    (clinic_id, patient_id, service_id, appointment_date,
                     start_time, end_time, status, full_name,
                     total_duration_minutes, notes)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                RETURNING id
                """,
                (
                    CLINIC_ID,
                    patient_id,
                    service_id,
                    date_key,
                    fromT,
                    toT,
                    "CONFIRMED",
                    patName.strip() if patName else None,
                    duration if duration > 0 else None,
                    notes,
                ),
            )
            appt_id = cur.fetchone()[0]
            inserted += 1
            print(f"  {date_key} {fromT}-{toT} | {patName} | patient_id={'found' if patient_id else 'MISSING'}")
        except Exception as e:
            errors.append(f"{patName} ({date_key} {fromT}): {e}")
            conn.rollback()

    conn.commit()
    cur.close()
    conn.close()

    print(f"\nInserted: {inserted}")
    print(f"Skipped: {skipped}")
    if errors:
        print(f"Errors ({len(errors)}):")
        for e in errors:
            print(f"  {e}")


if __name__ == "__main__":
    main()
